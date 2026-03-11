from __future__ import annotations

from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import PatternFill

from .helpers import set_auto_width_all
from .labels import AUTHOR_STAT_LABELS, RECORD_HEADER_LABELS, SERIES_NON_DUP, STAT_LABELS, STRICT, SUSPECT


def export_excel(
    rows: list[dict],
    headers: list[str],
    stats: dict,
    author_stats: list[dict],
    output_dir: Path,
) -> Path:
    wb = Workbook()

    ws_records = wb.active
    ws_records.title = "作品归档"
    if headers:
        ws_records.append([RECORD_HEADER_LABELS.get(h, h) for h in headers])
        for row in rows:
            ws_records.append([row.get(h, "") for h in headers])

        header_index = {h: i + 1 for i, h in enumerate(headers)}
        strict_fill = PatternFill(fill_type="solid", fgColor="FDE2E1")
        suspect_fill = PatternFill(fill_type="solid", fgColor="FFF4CC")
        series_fill = PatternFill(fill_type="solid", fgColor="E7F0FF")

        status_col = header_index.get("duplicate_status")
        if status_col is not None:
            for row_idx in range(2, ws_records.max_row + 1):
                status = str(ws_records.cell(row_idx, status_col).value or "")
                fill = None
                if status == STRICT:
                    fill = strict_fill
                elif status == SUSPECT:
                    fill = suspect_fill
                elif status == SERIES_NON_DUP:
                    fill = series_fill
                if fill:
                    for col_idx in range(1, ws_records.max_column + 1):
                        ws_records.cell(row_idx, col_idx).fill = fill

        set_auto_width_all(ws_records)

    ws_stats = wb.create_sheet("统计")
    ws_stats.append(["指标", "数值"])
    for key, value in stats.items():
        ws_stats.append([STAT_LABELS.get(key, key), value])
    set_auto_width_all(ws_stats)

    ws_author = wb.create_sheet("作者统计")
    ws_author.append(
        [
            AUTHOR_STAT_LABELS["author"],
            AUTHOR_STAT_LABELS["works"],
            AUTHOR_STAT_LABELS["strict_duplicates"],
            AUTHOR_STAT_LABELS["suspected_duplicates"],
            AUTHOR_STAT_LABELS["series_related_non_duplicates"],
            AUTHOR_STAT_LABELS["series_missing_items"],
            AUTHOR_STAT_LABELS["non_duplicates"],
            AUTHOR_STAT_LABELS["manual_review_items"],
            AUTHOR_STAT_LABELS.get("high_risk_items", "高风险待审"),
        ]
    )
    for row in author_stats:
        ws_author.append(
            [
                row["author"],
                row["works"],
                row["strict_duplicates"],
                row["suspected_duplicates"],
                row["series_related_non_duplicates"],
                row["series_missing_items"],
                row["non_duplicates"],
                row["manual_review_items"],
                row.get("high_risk_items", 0),
            ]
        )
    set_auto_width_all(ws_author)

    excel_path = output_dir / "作品归档结果.xlsx"
    try:
        wb.save(excel_path)
        return excel_path
    except PermissionError:
        fallback = output_dir / f"作品归档结果_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        wb.save(fallback)
        print(f"[Export][WARN] Excel 文件被占用，已另存为: {fallback}")
        return fallback
